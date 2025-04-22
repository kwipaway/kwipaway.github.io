import openpyxl
import re
import os
from glob import glob
from datetime import datetime
from tkinter import Tk, Label, Button, Text, END, filedialog, messagebox, Scrollbar, RIGHT, Y, BOTH

# Global list to hold all generated SQL statements.
global_sql_statements = []
# Set to track vendor IDs already processed (so we don't insert the same vendor twice).
vendors_seen = set()

# ------------------------------------------------------------------
# Helper Functions
# ------------------------------------------------------------------

def safe_sql_string(raw):
    """
    Wraps a string in single quotes and escapes internal quotes.
    Returns "NULL" if empty.
    """
    if raw is None:
        return "NULL"
    text = str(raw).strip()
    if text == "":
        return "NULL"
    safe_val = text.replace("'", "''")
    return f"'{safe_val}'"

def parse_due_date(due_date_cell):
    """
    Parses a string like "Due Date: 03/27/2025" or "03/27/2025" from cell D3.
    Returns only the date in YYYY-MM-DD format.
    If unparseable, returns None.
    """
    if not due_date_cell:
        return None
    text = str(due_date_cell).strip()
    # Remove any leading "Due Date:" text (case-insensitive)
    text = re.sub(r'(?i)^due date:\s*', '', text)
    try:
        dt = datetime.strptime(text, "%m/%d/%Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return None

def parse_invoice_date(cell_value):
    """
    Parses an invoice date from:
      - Python datetime (Excel date)
      - "YYYY-MM-DDT..."
      - "MM/DD/YYYY"
    Returns "YYYY-MM-DD" or None.
    """
    if not cell_value:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d")
    text = str(cell_value).strip()
    if 'T' in text:
        iso_part = text.split('T')[0]
        try:
            dt = datetime.strptime(iso_part, "%Y-%m-%d")
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        dt = datetime.strptime(text, "%m/%d/%Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return None

def convert_to_decimal_or_null(value):
    """
    Converts a value to a decimal string with 2 decimals; else returns "NULL".
    """
    if value is None or str(value).strip() == "":
        return "NULL"
    try:
        return f"{float(value):.2f}"
    except ValueError:
        return "NULL"

# ------------------------------------------------------------------
# Core Function: Generate SQL Inserts for a Single APV File
# ------------------------------------------------------------------

def generate_inserts_for_apv(excel_file_path):
    """
    Reads the APV Excel file and produces SQL INSERTs for:
      1. [dbo].[Vendor] (from A2, if not already inserted)
      2. [dbo].[AccountsPayableVoucher] for each invoice row,
         including:
           - VendorID, Invoice details, Account Number, 
           - Due Date in D3,
           - CompletedBy (C17), Comments (C18), AuthorizedBy (C20), PostedBy (C21).
    
    Assumes:
      - A2 => Vendor ("CRM")
      - D2 => Account Number ("Wood")
      - D3 => "Due Date: 03/27/2025" or "03/27/2025" (extract date)
      - Invoice data starts row 5
      - CompletedBy => C17
      - Comments => C18
      - AuthorizedBy => C20
      - PostedBy => C21
    Adjust row/col references as needed.
    """
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    statements = []

    # (A) Vendor from A2
    vendor_val = ws.cell(row=2, column=1).value
    vendor_id = vendor_val.strip() if vendor_val else ""
    vendor_name = vendor_id  
    if vendor_id and vendor_id not in vendors_seen:
        vendor_sql = (
            "INSERT INTO [dbo].[Vendor] ([VendorID], [Name]) VALUES ('"
            + vendor_id.replace("'", "''")
            + "', '"
            + vendor_name.replace("'", "''")
            + "');"
        )
        statements.append("-- Vendor Insert")
        statements.append(vendor_sql)
        vendors_seen.add(vendor_id)

    # (B) Account Number from D2
    acct_number_cell = ws.cell(row=2, column=4).value
    if acct_number_cell:
        acct_number_str = acct_number_cell.strip()
        acct_number_sql = safe_sql_string(acct_number_str)
    else:
        acct_number_sql = "NULL"

    # (C) Due Date from D3
    due_date_cell = ws.cell(row=3, column=4).value
    due_date_str = parse_due_date(due_date_cell)
    due_date_sql = safe_sql_string(due_date_str) if due_date_str else "NULL"

    # (D) Summary Fields from column C
    completed_by_cell = ws.cell(row=17, column=3).value  # C17
    completed_by_sql  = safe_sql_string(completed_by_cell)

    comments_cell = ws.cell(row=18, column=3).value  # C18
    comments_sql = safe_sql_string(comments_cell)

    authorized_by_cell = ws.cell(row=20, column=3).value  # C20
    authorized_by_sql = safe_sql_string(authorized_by_cell)

    posted_by_cell = ws.cell(row=21, column=3).value  # C21
    posted_by_sql = safe_sql_string(posted_by_cell)

    # (E) Invoice Line Items: row 5 onward
    start_row = 5
    max_row = ws.max_row

    statements.append("\n-- AccountsPayableVoucher Inserts")
    for row_idx in range(start_row, max_row + 1):
        invoice_date_val = ws.cell(row=row_idx, column=1).value
        if not invoice_date_val or str(invoice_date_val).strip() == "":
            break

        inv_date_str = parse_invoice_date(invoice_date_val)
        invoice_date_sql = safe_sql_string(inv_date_str)

        invoice_number_val = ws.cell(row=row_idx, column=2).value
        invoice_number_sql = safe_sql_string(invoice_number_val)

        gl_account_val = ws.cell(row=row_idx, column=3).value
        gl_account_sql = safe_sql_string(gl_account_val)

        description_val = ws.cell(row=row_idx, column=4).value
        description_sql = safe_sql_string(description_val)

        amount_val = ws.cell(row=row_idx, column=5).value
        amount_sql = convert_to_decimal_or_null(amount_val)

        voucher_sql = (
            "INSERT INTO [dbo].[AccountsPayableVoucher]\n"
            "       ([VendorID], [InvoiceNumber], [InvoiceDate], [DueDate], [Amount],\n"
            "        [AccountNumber], [GeneralLedgerAccount], [Description],\n"
            "        [CompletedBy], [Comments], [AuthorizedBy], [PostedBy])\n"
            "VALUES ("
            + safe_sql_string(vendor_id) + ", "
            + invoice_number_sql + ", "
            + invoice_date_sql + ", "
            + due_date_sql + ", "
            + amount_sql + ", "
            + acct_number_sql + ", "
            + gl_account_sql + ", "
            + description_sql + ", "
            + completed_by_sql + ", "
            + comments_sql + ", "
            + authorized_by_sql + ", "
            + posted_by_sql
            + ");"
        )
        statements.append(voucher_sql)

    return statements

# ------------------------------------------------------------------
# Folder Processing
# ------------------------------------------------------------------

def process_folder(folder_path):
    all_statements = []
    pattern = os.path.join(folder_path, "*.xlsx")
    file_list = glob(pattern)
    file_list.extend(glob(os.path.join(folder_path, "*.xls")))

    for file_path in file_list:
        try:
            stmts = generate_inserts_for_apv(file_path)
            header = f"-- Processing File: {file_path}"
            all_statements.append(header)
            all_statements.extend(stmts)
            all_statements.append("")
        except Exception as ex:
            all_statements.append(f"-- Error processing file {file_path}: {ex}")

    return all_statements

# ------------------------------------------------------------------
# Tkinter GUI Functions
# ------------------------------------------------------------------

def browse_excel_file(text_widget):
    file_path = filedialog.askopenfilename(
        title="Select an APV Excel Document",
        filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
    )
    if not file_path:
        return

    text_widget.delete(1.0, END)
    try:
        statements = generate_inserts_for_apv(file_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error processing file:\n{e}")
        return

    global global_sql_statements
    global_sql_statements.clear()

    header_line = f"-- Processing: {file_path}"
    text_widget.insert(END, header_line + "\n\n")
    global_sql_statements.append(header_line)
    for stmt in statements:
        text_widget.insert(END, stmt + "\n\n")
        global_sql_statements.append(stmt)

def browse_excel_folder(text_widget):
    folder_path = filedialog.askdirectory(title="Select a Folder with APV Documents")
    if not folder_path:
        return

    text_widget.delete(1.0, END)
    try:
        statements = process_folder(folder_path)
    except Exception as e:
        messagebox.showerror("Error", f"Error processing folder:\n{e}")
        return

    global global_sql_statements
    global_sql_statements.clear()

    header_line = f"-- Processing Folder: {folder_path}"
    text_widget.insert(END, header_line + "\n\n")
    global_sql_statements.append(header_line)
    for stmt in statements:
        text_widget.insert(END, stmt + "\n\n")
        global_sql_statements.append(stmt)

def save_sql_file():
    if not global_sql_statements:
        messagebox.showwarning("Warning", "No SQL statements to save.")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".sql",
        filetypes=[("SQL Files", "*.sql"), ("All Files", "*.*")],
        title="Save SQL File"
    )
    if not file_path:
        return

    try:
        with open(file_path, "w", encoding="utf-8") as f:
            for stmt in global_sql_statements:
                f.write(stmt + "\n\n" if stmt else "\n")
        messagebox.showinfo("Success", f"SQL statements saved to:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save SQL file:\n{e}")

def main():
    root = Tk()
    root.title("APV & Vendor SQL Generator - 4 Fields in Column C")
    root.geometry("900x700")

    label = Label(root, text="Select an APV Excel file or folder to process:")
    label.pack(pady=10)

    scrollbar = Scrollbar(root)
    scrollbar.pack(side=RIGHT, fill=Y)

    text_area = Text(root, wrap="none", yscrollcommand=scrollbar.set)
    text_area.pack(padx=10, pady=10, fill=BOTH, expand=True)
    scrollbar.config(command=text_area.yview)

    btn_frame = ButtonFrame(root, text_area)
    btn_frame.pack(pady=5)

    root.mainloop()

class ButtonFrame(Label):
    def __init__(self, master, text_area):
        super().__init__(master)
        self.text_area = text_area

        # Single file
        self.browse_file_btn = Button(
            master, text="Browse File",
            command=lambda: browse_excel_file(text_area)
        )
        self.browse_file_btn.pack(pady=5)

        # Entire folder
        self.browse_folder_btn = Button(
            master, text="Browse Folder",
            command=lambda: browse_excel_folder(text_area)
        )
        self.browse_folder_btn.pack(pady=5)

        # Save to .sql
        self.save_btn = Button(master, text="Save to .sql", command=save_sql_file)
        self.save_btn.pack(pady=5)

if __name__ == "__main__":
    main()
